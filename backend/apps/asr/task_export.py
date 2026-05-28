"""测试任务结果导出为 Excel（WER / SDI / 数据集汇总 / 音频明细）."""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

from .k2_engine.lang_mapping import platform_lang_to_k2_scp_lang
from .models import TestAudioResult, TestTask, TestTaskDataset

# 与旧 K2 脚本 comm.py 中汇报矩阵语种顺序一致
LANGUAGE_ORDER = ["zh-cn", "en", "es", "ja", "ko", "th", "fr", "de", "it", "ar", "ru"]

NA = "N/A"


def safe_export_filename(task_name: str) -> str:
    cleaned = "".join(c if c.isalnum() or c in "-_" else "_" for c in (task_name or "task"))
    return (cleaned[:60] or "task") + "_results.xlsx"


def _pct(value: float) -> str:
    return f"{value * 100:.2f}%"


def _write_sheet_header(ws, headers: list[str]) -> None:
    for col, title in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=title)


def _build_wer_sdi_matrices(
    task_datasets: list[TestTaskDataset],
) -> tuple[list[str], dict[str, dict[str, str]], dict[str, dict[str, str]]]:
    dataset_names = [td.dataset.name for td in task_datasets]
    wer_matrix: dict[str, dict[str, str]] = {lang: {} for lang in LANGUAGE_ORDER}
    sdi_matrix: dict[str, dict[str, str]] = {lang: {} for lang in LANGUAGE_ORDER}

    for td in task_datasets:
        ds_name = td.dataset.name
        lang_row = platform_lang_to_k2_scp_lang(td.dataset.language)
        if td.total_audio > 0:
            wer_matrix.setdefault(lang_row, {})[ds_name] = _pct(td.avg_wer)
            sdi_matrix.setdefault(lang_row, {})[ds_name] = f"{td.s_cnt}/{td.d_cnt}/{td.i_cnt}"
        for lang in LANGUAGE_ORDER:
            wer_matrix[lang].setdefault(ds_name, NA)
            sdi_matrix[lang].setdefault(ds_name, NA)

    return dataset_names, wer_matrix, sdi_matrix


def _write_matrix_sheet(ws, title_col: str, dataset_names: list[str], matrix: dict[str, dict[str, str]]) -> None:
    _write_sheet_header(ws, [title_col, *dataset_names])
    for row_idx, lang in enumerate(LANGUAGE_ORDER, start=2):
        ws.cell(row=row_idx, column=1, value=lang)
        for col_idx, ds_name in enumerate(dataset_names, start=2):
            ws.cell(row=row_idx, column=col_idx, value=matrix[lang].get(ds_name, NA))


def build_task_export_xlsx(task: TestTask, task_datasets: list[TestTaskDataset]) -> BytesIO:
    """生成含 4 个 sheet 的结果 Excel，明细为单 sheet 合并."""
    wb = Workbook()
    wb.remove(wb.active)

    dataset_names, wer_matrix, sdi_matrix = _build_wer_sdi_matrices(task_datasets)

    ws_wer = wb.create_sheet("WER")
    _write_matrix_sheet(ws_wer, "Language", dataset_names, wer_matrix)

    ws_sdi = wb.create_sheet("SDI")
    _write_matrix_sheet(ws_sdi, "Language", dataset_names, sdi_matrix)

    ws_summary = wb.create_sheet("数据集汇总")
    summary_headers = [
        "数据集",
        "语种",
        "音频数",
        "总时长(秒)",
        "平均WER",
        "RET",
        "S",
        "I",
        "D",
        "Hit",
    ]
    _write_sheet_header(ws_summary, summary_headers)
    for row_idx, td in enumerate(task_datasets, start=2):
        ws_summary.cell(row=row_idx, column=1, value=td.dataset.name)
        ws_summary.cell(row=row_idx, column=2, value=td.dataset.language)
        ws_summary.cell(row=row_idx, column=3, value=td.total_audio)
        ws_summary.cell(row=row_idx, column=4, value=td.total_duration)
        ws_summary.cell(row=row_idx, column=5, value=_pct(td.avg_wer) if td.total_audio else NA)
        ws_summary.cell(row=row_idx, column=6, value=_pct(td.ret) if td.total_audio else NA)
        ws_summary.cell(row=row_idx, column=7, value=td.s_cnt)
        ws_summary.cell(row=row_idx, column=8, value=td.i_cnt)
        ws_summary.cell(row=row_idx, column=9, value=td.d_cnt)
        ws_summary.cell(row=row_idx, column=10, value=td.hit_cnt)

    ws_detail = wb.create_sheet("音频明细")
    detail_headers = [
        "数据集",
        "音频名称",
        "语种",
        "时长(秒)",
        "WER",
        "S",
        "I",
        "D",
        "参考文本",
        "识别文本",
    ]
    _write_sheet_header(ws_detail, detail_headers)

    results_qs = (
        TestAudioResult.objects.filter(task=task)
        .select_related("audio", "dataset")
        .order_by("dataset__name", "audio__name", "id")
    )
    row_idx = 2
    for result in results_qs.iterator(chunk_size=500):
        errors = result.errors_json or {}
        ws_detail.cell(row=row_idx, column=1, value=result.dataset.name)
        ws_detail.cell(row=row_idx, column=2, value=result.audio.name)
        ws_detail.cell(row=row_idx, column=3, value=result.audio.language)
        ws_detail.cell(row=row_idx, column=4, value=result.audio.duration or 0)
        ws_detail.cell(row=row_idx, column=5, value=_pct(result.wer))
        ws_detail.cell(row=row_idx, column=6, value=errors.get("s_cnt", 0))
        ws_detail.cell(row=row_idx, column=7, value=errors.get("i_cnt", 0))
        ws_detail.cell(row=row_idx, column=8, value=errors.get("d_cnt", 0))
        ws_detail.cell(row=row_idx, column=9, value=result.ref_text or "")
        ws_detail.cell(row=row_idx, column=10, value=result.hyp_text or "")
        row_idx += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
