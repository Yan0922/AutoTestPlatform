"""ASR 模块视图层."""
import os
from io import BytesIO

from django.conf import settings
from django.db import transaction
from django.db.models import Count, Q, Sum
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from rest_framework import status as drf_status
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from .audio_utils import detect_duration_from_path, resolve_industry
from .model_utils import build_model_dir_path
from .remote_download_job import (
    cancel_remote_download_job,
    get_job_status,
    start_remote_download_job,
)
from .remote_model import fetch_remote_catalog, get_remote_base_url
from .filters import AudioFilter
from .models import (
    Audio,
    AsrModel,
    AsrModelFile,
    Dataset,
    DatasetAudio,
    TestAudioResult,
    TestTask,
    TestTaskDataset,
)
from .serializers import (
    AsrModelSerializer,
    AudioSerializer,
    DatasetSerializer,
    TestAudioResultSerializer,
    TestTaskDatasetSerializer,
    TestTaskSerializer,
)
from .task_export import build_task_export_xlsx, safe_export_filename
from .task_run_group import (
    get_run_group_root,
    resolve_run_task,
    serialize_run_group,
)
from .task_utils import build_task_name, extract_task_base_name
from .tasks import start_test_task_async


class AudioPagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = "page_size"
    max_page_size = 200


class TaskListPagination(PageNumberPagination):
    page_size = 15
    page_size_query_param = "page_size"


class DatasetPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = "page_size"


# ============================
# 模型管理
# ============================
class AsrModelViewSet(viewsets.ModelViewSet):
    """ASR 模型 CRUD."""

    serializer_class = AsrModelSerializer
    queryset = AsrModel.objects.filter(status=1).prefetch_related("files")
    parser_classes = [JSONParser, MultiPartParser, FormParser]
    search_fields = ["name"]

    @staticmethod
    def _multi_values(query_params, key: str) -> list[str]:
        raw = query_params.getlist(key)
        if not raw:
            return []
        values: list[str] = []
        for item in raw:
            values.extend(p.strip() for p in str(item).split(",") if p.strip())
        return values

    def get_queryset(self):
        qs = AsrModel.objects.filter(status=1).prefetch_related("files")
        keyword = self.request.query_params.get("search") or self.request.query_params.get("name")
        if keyword:
            qs = qs.filter(name__icontains=keyword)
        languages = self._multi_values(self.request.query_params, "language")
        if languages:
            qs = qs.filter(language__in=languages)
        sizes = self._multi_values(self.request.query_params, "size")
        if sizes:
            qs = qs.filter(size__in=sizes)
        return qs.order_by("-created_at")

    def create(self, request, *args, **kwargs):
        data = {
            "name": request.data.get("name"),
            "language": request.data.get("language", "zh"),
            "version": request.data.get("version"),
            "size": request.data.get("size", "base"),
        }
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        with transaction.atomic():
            instance = serializer.save()
            instance.dir_path = build_model_dir_path(instance.name, instance.id)
            instance.save(update_fields=["dir_path"])
            self._save_files(instance, request.FILES.getlist("files"))
        return Response(self.get_serializer(instance).data, status=drf_status.HTTP_201_CREATED)

    @staticmethod
    def _save_files(model_instance: AsrModel, files):
        if not files:
            return
        base_dir = os.path.join(settings.MEDIA_ROOT, model_instance.dir_path)
        os.makedirs(base_dir, exist_ok=True)
        for f in files:
            file_name = getattr(f, "name", None) or f.field_name
            dest = os.path.join(base_dir, file_name)
            os.makedirs(os.path.dirname(dest), exist_ok=True)
            size = 0
            with open(dest, "wb") as out:
                for chunk in f.chunks():
                    out.write(chunk)
                    size += len(chunk)
            AsrModelFile.objects.create(
                model=model_instance,
                file_name=file_name,
                file_size=size,
                file_path=f"{model_instance.dir_path}/{file_name}",
            )

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        data = {
            "name": request.data.get("name", instance.name),
            "language": request.data.get("language", instance.language),
            "version": request.data.get("version", instance.version),
            "size": request.data.get("size", instance.size),
        }
        serializer = self.get_serializer(instance, data=data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.status = 0
        instance.save(update_fields=["status"])
        return Response(status=drf_status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=["get"], url_path="remote-config")
    def remote_config(self, request):
        """返回远程模型下载根地址配置."""
        return Response({"base_url": get_remote_base_url()})

    @action(detail=False, methods=["get"], url_path="remote-catalog")
    def remote_catalog(self, request):
        """根据语种+尺寸拉取远程目录中可用的模型版本."""
        languages = self._multi_values(request.query_params, "language")
        sizes = self._multi_values(request.query_params, "size")
        if not languages or not sizes:
            return Response({"detail": "请先选择语种和尺寸"}, status=400)
        try:
            data = fetch_remote_catalog(languages, sizes)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=400)
        except Exception as exc:  # noqa: BLE001
            return Response({"detail": f"拉取远程目录失败: {exc}"}, status=502)
        return Response(data)

    @action(detail=False, methods=["post"], url_path="remote-download/start")
    def remote_download_start(self, request):
        """后台启动远程模型下载，立即返回 job_id."""
        languages = request.data.get("languages") or []
        sizes = request.data.get("sizes") or []
        versions = request.data.get("versions") or []
        if not languages or not sizes or not versions:
            return Response({"detail": "languages / sizes / versions 均为必填"}, status=400)
        try:
            job_id = start_remote_download_job(languages, sizes, versions)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=400)
        status_data = get_job_status(job_id) or {"job_id": job_id, "status": "running"}
        return Response(status_data, status=drf_status.HTTP_202_ACCEPTED)

    @action(detail=False, methods=["get"], url_path="remote-download/status")
    def remote_download_status(self, request):
        """查询下载任务进度（不传 job_id 时返回当前/最近一次任务）."""
        job_id = request.query_params.get("job_id")
        data = get_job_status(job_id)
        if not data:
            return Response({"status": "idle"})
        return Response(data)

    @action(detail=False, methods=["post"], url_path="remote-download/cancel")
    def remote_download_cancel(self, request):
        """取消进行中的下载任务（仅「取消」按钮应调用）."""
        job_id = request.data.get("job_id")
        if not job_id:
            return Response({"detail": "job_id 必填"}, status=400)
        if not cancel_remote_download_job(job_id):
            return Response({"detail": "任务不存在或已结束"}, status=400)
        return Response(get_job_status(job_id) or {"job_id": job_id, "status": "cancelled"})


# ============================
# 数据集管理
# ============================
class DatasetViewSet(viewsets.ModelViewSet):
    serializer_class = DatasetSerializer
    pagination_class = DatasetPagination

    def get_queryset(self):
        qs = Dataset.objects.filter(status=1).annotate(
            total_audio=Count("audios", filter=Q(audios__status=1)),
            total_duration=Sum("audios__duration", filter=Q(audios__status=1)),
        )
        keyword = self.request.query_params.get("search") or self.request.query_params.get("name")
        if keyword:
            qs = qs.filter(name__icontains=keyword)
        return qs.order_by("-created_at")

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        with transaction.atomic():
            DatasetAudio.objects.filter(dataset=instance).delete()
            instance.status = 0
            instance.save(update_fields=["status"])
        return Response(status=drf_status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=["post"], url_path="remove-audios")
    def remove_audios(self, request, pk=None):
        """从数据集中移除指定音频."""
        dataset = self.get_object()
        audio_ids = request.data.get("audio_ids", [])
        DatasetAudio.objects.filter(dataset=dataset, audio_id__in=audio_ids).delete()
        return Response({"removed": len(audio_ids)})


# ============================
# 数据池 / 音频管理
# ============================
class AudioViewSet(viewsets.ModelViewSet):
    serializer_class = AudioSerializer
    filterset_class = AudioFilter
    pagination_class = AudioPagination

    def get_queryset(self):
        qs = Audio.objects.filter(status=1).prefetch_related("datasets")
        return qs.order_by("-created_at")

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.status = 0
        instance.save(update_fields=["status"])
        return Response(status=drf_status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=["post"], url_path="batch-delete")
    def batch_delete(self, request):
        ids = request.data.get("ids", [])
        Audio.objects.filter(id__in=ids).update(status=0)
        return Response({"deleted": len(ids)})

    @action(detail=False, methods=["post"], url_path="join-dataset")
    def join_dataset(self, request):
        """将音频加入到一个或多个数据集."""
        audio_ids = request.data.get("audio_ids", [])
        dataset_ids = request.data.get("dataset_ids", [])
        records = []
        for aid in audio_ids:
            for did in dataset_ids:
                records.append(DatasetAudio(audio_id=aid, dataset_id=did))
        DatasetAudio.objects.bulk_create(records, ignore_conflicts=True)
        return Response({"count": len(records)})

    @action(detail=True, methods=["patch"], url_path="update-text")
    def update_text(self, request, pk=None):
        instance = self.get_object()
        new_text = request.data.get("ref_text", "")
        instance.ref_text = new_text
        instance.save(update_fields=["ref_text"])
        return Response(AudioSerializer(instance).data)

    @action(detail=True, methods=["patch"], url_path="update-info")
    def update_info(self, request, pk=None):
        """更新音频名称、来源、噪声、行业及关联数据集."""
        instance = self.get_object()
        allowed = ("name", "source", "noise", "industry")
        update_fields = []
        for field in allowed:
            if field in request.data:
                setattr(instance, field, request.data[field])
                update_fields.append(field)
        if update_fields:
            instance.save(update_fields=update_fields)

        if "dataset_ids" in request.data:
            dataset_ids = request.data.get("dataset_ids") or []
            with transaction.atomic():
                DatasetAudio.objects.filter(audio=instance).delete()
                if dataset_ids:
                    DatasetAudio.objects.bulk_create([
                        DatasetAudio(audio=instance, dataset_id=did) for did in dataset_ids
                    ])

        instance.refresh_from_db()
        return Response(AudioSerializer(instance).data)

    @action(detail=False, methods=["get"], url_path="template")
    def template(self, request):
        """下载导入模板."""
        wb = Workbook()
        ws = wb.active
        ws.title = "audio_template"
        headers = ["音频名称", "语种", "音频路径", "音频来源", "噪声", "行业", "文本"]
        ws.append(headers)
        ws.append(["示例音频001.wav", "zh", "/media/audio/example.wav", "Sota1", "quiet", "technology", "你好世界"])
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        resp = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="audio_template.xlsx"'
        return resp

    @staticmethod
    def _parse_excel_row(headers: list[str], row: tuple) -> dict:
        """按表头解析一行 Excel 数据，时长由系统自动识别."""
        col_map = {str(h or "").strip(): i for i, h in enumerate(headers)}

        def cell(key: str, default=""):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return default
            val = row[idx]
            return "" if val is None else val

        audio_path = str(cell("音频路径")).strip()
        return {
            "name": str(cell("音频名称")).strip(),
            "language": str(cell("语种") or "zh").strip() or "zh",
            "audio_path": audio_path,
            "source": str(cell("音频来源") or "outside").strip() or "outside",
            "noise": str(cell("噪声") or "quiet").strip() or "quiet",
            "industry": resolve_industry(str(cell("行业") or "unknown")),
            "ref_text": str(cell("文本")).strip(),
            "duration": detect_duration_from_path(audio_path),
        }

    @action(detail=False, methods=["post"], url_path="parse-excel", parser_classes=[MultiPartParser])
    def parse_excel(self, request):
        """解析上传的 Excel，返回预览数据，不入库."""
        f = request.FILES.get("file")
        if not f:
            return Response({"detail": "请上传文件"}, status=400)
        try:
            wb = load_workbook(filename=BytesIO(f.read()), data_only=True)
            ws = wb.active
        except Exception as exc:  # noqa: BLE001
            return Response({"detail": f"文件解析失败: {exc}"}, status=400)
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({"headers": [], "rows": []})
        headers = [str(c or "") for c in rows[0]]
        data_rows = []
        for row in rows[1:]:
            if not any(row):
                continue
            data_rows.append(self._parse_excel_row(headers, row))
        return Response({"headers": headers, "rows": data_rows})

    @action(detail=False, methods=["post"], url_path="import")
    def import_data(self, request):
        rows = request.data.get("rows", [])
        objs = []
        for r in rows:
            audio_path = r.get("audio_path") or ""
            duration = r.get("duration")
            if duration in (None, "", 0):
                duration = detect_duration_from_path(audio_path)
            objs.append(Audio(
                name=r.get("name") or "未命名",
                language=r.get("language") or "zh",
                audio_path=audio_path,
                source=r.get("source") or "outside",
                noise=r.get("noise") or "quiet",
                industry=resolve_industry(r.get("industry") or "unknown"),
                duration=float(duration or 0),
                ref_text=r.get("ref_text") or "",
            ))
        created = Audio.objects.bulk_create(objs)
        return Response({"count": len(created)})


class AudioUploadView(APIView):
    """单个音频文件上传接口，返回服务端可访问的 URL."""

    parser_classes = [MultiPartParser]

    def post(self, request):
        f = request.FILES.get("file")
        if not f:
            return Response({"detail": "请上传文件"}, status=400)
        sub = "audio"
        base_dir = os.path.join(settings.MEDIA_ROOT, sub)
        os.makedirs(base_dir, exist_ok=True)
        dest = os.path.join(base_dir, f.name)
        size = 0
        with open(dest, "wb") as out:
            for chunk in f.chunks():
                out.write(chunk)
                size += len(chunk)
        return Response({
            "name": f.name,
            "size": size,
            "audio_path": f"{settings.MEDIA_URL}{sub}/{f.name}",
        })


# ============================
# 测试任务
# ============================
class TestTaskViewSet(viewsets.ModelViewSet):
    serializer_class = TestTaskSerializer
    pagination_class = TaskListPagination

    def get_queryset(self):
        qs = TestTask.objects.filter(status=1).select_related("model")
        if self.action in ("result", "export", "rerun", "retrieve"):
            return qs.order_by("-created_at")
        return qs.filter(root_task__isnull=True).order_by("-created_at")

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.status = 0
        instance.save(update_fields=["status"])
        return Response(status=drf_status.HTTP_204_NO_CONTENT)

    def create(self, request, *args, **kwargs):
        name = request.data.get("name")
        model_id = request.data.get("model")
        dataset_ids = request.data.get("dataset_ids", [])
        raw_name = (name or "").strip()
        if not raw_name or not model_id or not dataset_ids:
            return Response({"detail": "name / model / dataset_ids 都是必填"}, status=400)

        final_name = build_task_name(raw_name)

        with transaction.atomic():
            task = TestTask.objects.create(name=final_name, model_id=model_id)
            for did in dataset_ids:
                TestTaskDataset.objects.create(task=task, dataset_id=did)

        start_test_task_async(task.id)
        return Response(self.get_serializer(task).data, status=drf_status.HTTP_202_ACCEPTED)

    @action(detail=True, methods=["get"], url_path="result")
    def result(self, request, pk=None):
        """获取任务的统计与音频结果列表（支持分页）."""
        anchor = self.get_object()
        root = get_run_group_root(anchor)

        run_id_raw = request.query_params.get("run_id")
        run_id = None
        if run_id_raw:
            try:
                run_id = int(run_id_raw)
            except (TypeError, ValueError):
                run_id = None

        task = resolve_run_task(root, run_id)
        dataset_id = request.query_params.get("dataset_id")
        task_datasets = TestTaskDataset.objects.filter(task=task).select_related("dataset")
        td_data = TestTaskDatasetSerializer(task_datasets, many=True).data

        if dataset_id:
            target_dataset_id = int(dataset_id)
        else:
            first = task_datasets.first()
            target_dataset_id = first.dataset_id if first else None

        results_qs = (
            TestAudioResult.objects.filter(task=task, dataset_id=target_dataset_id)
            .select_related("audio")
            .order_by("audio__name", "id")
            if target_dataset_id
            else TestAudioResult.objects.none()
        )

        try:
            page = max(1, int(request.query_params.get("page", 1)))
            page_size = min(max(1, int(request.query_params.get("page_size", 20))), 100)
        except (TypeError, ValueError):
            page, page_size = 1, 20

        total = results_qs.count()
        start = (page - 1) * page_size
        results = results_qs[start : start + page_size]
        result_data = TestAudioResultSerializer(results, many=True).data
        root_data = TestTaskSerializer(root).data
        return Response({
            "task": TestTaskSerializer(task).data,
            "root_task": root_data,
            "run_group": serialize_run_group(root, current_run_id=task.id),
            "datasets": td_data,
            "audio_results": result_data,
            "audio_results_count": total,
            "page": page,
            "page_size": page_size,
        })

    @action(detail=True, methods=["get"], url_path="export")
    def export_results(self, request, pk=None):
        """导出任务结果为 Excel（WER / SDI / 数据集汇总 / 音频明细）."""
        anchor = self.get_object()
        root = get_run_group_root(anchor)

        run_id_raw = request.query_params.get("run_id")
        run_id = None
        if run_id_raw:
            try:
                run_id = int(run_id_raw)
            except (TypeError, ValueError):
                run_id = None
        task = resolve_run_task(root, run_id)

        if task.task_status == 1:
            return Response({"detail": "任务进行中，请完成后导出"}, status=400)

        task_datasets = list(
            TestTaskDataset.objects.filter(task=task)
            .select_related("dataset")
            .order_by("dataset__name")
        )
        has_results = any(td.total_audio > 0 for td in task_datasets) or TestAudioResult.objects.filter(
            task=task
        ).exists()
        if not has_results:
            return Response({"detail": "暂无结果可导出"}, status=400)

        buf = build_task_export_xlsx(task, task_datasets)
        filename = safe_export_filename(task.name)
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @action(detail=True, methods=["post"], url_path="rerun")
    def rerun(self, request, pk=None):
        """按原任务配置创建新任务并后台运行（不覆盖原任务结果）."""
        source = (
            TestTask.objects.filter(status=1, pk=self.get_object().pk)
            .select_related("model")
            .prefetch_related("task_datasets__dataset")
            .first()
        )
        if source is None:
            return Response({"detail": "任务不存在"}, status=404)

        root = get_run_group_root(source)

        if root.model.status != 1:
            return Response(
                {"detail": f"模型「{root.model.name}」已删除，无法重新运行"},
                status=400,
            )

        dataset_ids: list[int] = []
        deleted_names: list[str] = []
        for td in root.task_datasets.all():
            if td.dataset.status != 1:
                deleted_names.append(td.dataset.name)
            else:
                dataset_ids.append(td.dataset_id)

        if deleted_names:
            return Response(
                {"detail": f"数据集已删除，无法重新运行：{', '.join(deleted_names)}"},
                status=400,
            )
        if not dataset_ids:
            return Response({"detail": "原任务未关联有效数据集"}, status=400)

        root_id = root.id
        final_name = build_task_name(extract_task_base_name(root.name))
        with transaction.atomic():
            task = TestTask.objects.create(
                name=final_name,
                model_id=root.model_id,
                root_task_id=root_id,
            )
            for did in dataset_ids:
                TestTaskDataset.objects.create(task=task, dataset_id=did)

        start_test_task_async(task.id)
        root = TestTask.objects.get(pk=root_id)
        return Response(self.get_serializer(root).data, status=drf_status.HTTP_202_ACCEPTED)
